#!/usr/bin/env python3
"""
Fetch EU Weekly Oil Bulletin — heating gas oil prices + pass-through analysis.

Downloads the historical XLSX published every Thursday by the European Commission,
extracts:
  - Ireland weekly heating gas oil prices, with-tax and pre-tax (EUR/litre, 2yr)
  - Latest EU member-state comparison (with-tax)
  - Gross supply-chain margin history (pre-tax retail minus Brent crude cost)
  - Asymmetric pass-through statistics (β_up vs β_down)

Reads data/prices.json (written first by fetch_prices.py) for Brent/EUR/USD data.
Saves results to data/eu_bulletin.json.
No API key required.
"""

import io
import json
import os
import re
import sys
from datetime import datetime, timezone, timedelta

import requests
from bs4 import BeautifulSoup
import openpyxl

# ── Config ────────────────────────────────────────────────────────────────────

SCRIPT_DIR   = os.path.dirname(os.path.abspath(__file__))
OUTPUT_PATH  = os.path.join(SCRIPT_DIR, "..", "data", "eu_bulletin.json")
PRICES_PATH  = os.path.join(SCRIPT_DIR, "..", "data", "prices.json")

BULLETIN_PAGE = "https://energy.ec.europa.eu/data-and-analysis/weekly-oil-bulletin_en"
BASE_URL      = "https://energy.ec.europa.eu"

HISTORY_WEEKS  = 104   # ~2 years
LITRES_PER_BBL = 158.987

# ── Ireland Carbon Tax on Home Heating (kerosene / heating gas oil) ───────────
# Rate changes effective 1 May each year (Budget announcement, applied May)
CARBON_TAX_SCHEDULE = [
    ("2026-05-01", 71.00),
    ("2025-05-01", 63.50),
    ("2024-05-01", 56.00),
    ("2023-05-01", 48.50),
    ("2022-05-01", 41.00),
    ("2021-05-01", 33.50),
]
CO2_FACTOR_KEROSENE = 0.0025407  # tonnes CO2 per litre of kerosene/heating gas oil


def carbon_rate_for_week(week_str: str) -> float:
    """Return the carbon tax rate (€/tonne CO2) applicable for a given week (YYYY-MM-DD)."""
    for effective_date, rate in CARBON_TAX_SCHEDULE:
        if week_str >= effective_date:
            return rate
    return CARBON_TAX_SCHEDULE[-1][1]  # earliest known rate

COUNTRY_NAMES = {
    "AT": "Austria",     "BE": "Belgium",     "BG": "Bulgaria",
    "CY": "Cyprus",      "CZ": "Czech Rep.",  "DE": "Germany",
    "DK": "Denmark",     "EE": "Estonia",     "ES": "Spain",
    "FI": "Finland",     "FR": "France",      "GR": "Greece",
    "HR": "Croatia",     "HU": "Hungary",     "IE": "Ireland",
    "IT": "Italy",       "LT": "Lithuania",   "LU": "Luxembourg",
    "LV": "Latvia",      "MT": "Malta",       "NL": "Netherlands",
    "PL": "Poland",      "PT": "Portugal",    "RO": "Romania",
    "SE": "Sweden",      "SI": "Slovenia",    "SK": "Slovakia",
}

HEADERS = {"User-Agent": "Mozilla/5.0 (compatible; oil-dashboard-bot/1.0)"}

# ── Fetch / download ──────────────────────────────────────────────────────────

def find_xlsx_url() -> str:
    resp = requests.get(BULLETIN_PAGE, timeout=30, headers=HEADERS)
    resp.raise_for_status()
    soup = BeautifulSoup(resp.text, "lxml")
    for a in soup.find_all("a", href=True):
        href = a["href"]
        if "History" in href and href.lower().endswith(".xlsx"):
            return BASE_URL + href if href.startswith("/") else href
    raise RuntimeError("Could not find historical XLSX link on EU Oil Bulletin page")


def download_xlsx(url: str) -> bytes:
    resp = requests.get(url, timeout=120, headers=HEADERS)
    resp.raise_for_status()
    return resp.content


# ── XLSX parsing ──────────────────────────────────────────────────────────────

def _ie_col(headers: list, pattern: str) -> int | None:
    for i, h in enumerate(headers):
        if re.search(pattern, h, re.IGNORECASE):
            return i
    return None


def _date_str(val) -> str | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    s = str(val)[:10]
    return s if len(s) == 10 else None


def parse_with_tax(data: bytes) -> tuple[list, dict, str | None]:
    """
    Returns:
      - ireland_history: [{week, price_per_litre}, ...] newest first
      - eu_comparison: [{code, name, price_per_litre, is_ireland}, ...]
      - first_valid_date: str | None
    """
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    # Find sheet
    sheet_name = next((n for n in wb.sheetnames if "tax" in n.lower() and "wo" not in n.lower()), wb.sheetnames[0])
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise RuntimeError(f"Sheet '{sheet_name}' has only {len(rows)} rows")

    headers = [str(h) if h else "" for h in rows[0]]
    ie_col  = _ie_col(headers, r"IE_price_with_tax_heIEing_oil")
    if ie_col is None:
        raise RuntimeError("No Ireland with-tax column found")

    country_cols: dict[str, int] = {}
    for i, h in enumerate(headers):
        m = re.match(r"^([A-Z]{2})_price_with_tax_he.*ing_oil", h, re.I)
        if m:
            code = m.group(1).upper()
            if code in COUNTRY_NAMES:
                country_cols[code] = i

    ireland_history: list[dict]  = []
    latest_eu_prices: dict       = {}
    first_valid_date: str | None = None

    for row in rows[3:]:
        d = _date_str(row[0])
        if not d:
            continue
        v = row[ie_col] if ie_col < len(row) else None
        if v is None:
            continue
        try:
            price = round(float(v) / 1000.0, 4)
        except (ValueError, TypeError):
            continue
        ireland_history.append({"week": d, "price_per_litre": price})
        if first_valid_date is None:
            first_valid_date = d
            for code, col in country_cols.items():
                cv = row[col] if col < len(row) else None
                if cv is not None:
                    try:
                        latest_eu_prices[code] = round(float(cv) / 1000.0, 4)
                    except (ValueError, TypeError):
                        pass
        if len(ireland_history) >= HISTORY_WEEKS:
            break

    eu_comparison = sorted(
        [{"code": c, "name": COUNTRY_NAMES.get(c, c), "price_per_litre": p, "is_ireland": c == "IE"}
         for c, p in latest_eu_prices.items()],
        key=lambda x: x["price_per_litre"],
    )
    return ireland_history, eu_comparison, first_valid_date


def parse_pretax(data: bytes) -> list:
    """Extract Ireland pre-tax heating oil history from 'Prices wo taxes' sheet."""
    wb = openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    sheet_name = next((n for n in wb.sheetnames if "wo" in n.lower()), None)
    if sheet_name is None:
        raise RuntimeError("'Prices wo taxes' sheet not found")
    ws   = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 4:
        raise RuntimeError("Pre-tax sheet too small")

    headers = [str(h) if h else "" for h in rows[0]]
    ie_col  = _ie_col(headers, r"IE_price_wo_tax_heIEing_oil")
    if ie_col is None:
        raise RuntimeError("No Ireland pre-tax column found")

    history = []
    for row in rows[3:]:
        d = _date_str(row[0])
        if not d:
            continue
        v = row[ie_col] if ie_col < len(row) else None
        if v is None:
            continue
        try:
            history.append({"week": d, "pretax_eur_l": round(float(v) / 1000.0, 4)})
        except (ValueError, TypeError):
            pass
        if len(history) >= HISTORY_WEEKS:
            break
    return history


# ── Margin + pass-through computation ────────────────────────────────────────

def _nearest(lookup: dict, target: str, max_days: int = 7):
    """Find value in {date_str: value} lookup nearest to target date string."""
    if not lookup:
        return None
    try:
        td = datetime.strptime(target, "%Y-%m-%d")
    except ValueError:
        return None
    closest = min(lookup.keys(), key=lambda d: abs((datetime.strptime(d, "%Y-%m-%d") - td).days))
    if abs((datetime.strptime(closest, "%Y-%m-%d") - td).days) <= max_days:
        return lookup[closest]
    return None


def compute_margin_history(
    pretax_history: list,
    withtax_history: list,
    brent_weekly: list,
    eurusd_daily: list,
) -> list:
    """
    For each EU Bulletin week, compute:
      - pretax_eur_l:   Ireland ex-tax retail price (EU Bulletin)
      - withtax_eur_l:  Ireland with-tax retail price (EU Bulletin)
      - brent_eur_l:    Brent crude converted to EUR/litre
      - gross_margin:   pretax - brent  (supply chain markup, excl. tax)
      - tax_eur_l:      withtax - pretax
    """
    brent_map  = {b["period"]: b["value"] for b in brent_weekly}
    eurusd_map = {e["date"]:   e["rate"]  for e in eurusd_daily}
    withtax_map = {h["week"]: h["price_per_litre"] for h in withtax_history}

    result = []
    for h in pretax_history:
        wk      = h["week"]
        pretax  = h["pretax_eur_l"]
        withtax = withtax_map.get(wk)

        brent_usd = _nearest(brent_map,  wk, max_days=7)
        eurusd    = _nearest(eurusd_map, wk, max_days=5)

        if brent_usd is None or eurusd is None:
            continue

        brent_eur_l  = round(brent_usd / (eurusd * LITRES_PER_BBL), 4)
        gross_margin = round(pretax - brent_eur_l, 4)
        tax_eur_l    = round(withtax - pretax, 4) if withtax is not None else None

        # Tax decomposition
        carbon_eur_l    = round(carbon_rate_for_week(wk) * CO2_FACTOR_KEROSENE, 4)
        vat_eur_l       = round(withtax * (0.135 / 1.135), 4) if withtax is not None else None
        # excise = withtax/1.135 - pretax; other_mot = excise - carbon
        other_mot_eur_l = round((withtax / 1.135 - pretax) - carbon_eur_l, 4) if withtax is not None else None

        result.append({
            "week":           wk,
            "pretax_eur_l":   pretax,
            "withtax_eur_l":  withtax,
            "brent_eur_l":    brent_eur_l,
            "gross_margin":   gross_margin,
            "tax_eur_l":      tax_eur_l,
            "carbon_eur_l":   carbon_eur_l,
            "vat_eur_l":      vat_eur_l,
            "other_mot_eur_l": other_mot_eur_l,
        })

    # Return newest first (match existing convention)
    result.sort(key=lambda x: x["week"], reverse=True)
    return result


def _pass_through_core(chron: list, threshold: float = 0.003, lag: int = 0) -> dict | None:
    """
    Core pass-through computation for a given lag (weeks).
    lag=0 means same-week; lag=1 means retail responds one week after Brent moves.
    Returns None if insufficient data.
    """
    if len(chron) < max(8, lag + 2):
        return None

    d_retail = []
    d_brent  = []
    for i in range(1 + lag, len(chron)):
        dr = chron[i]["pretax_eur_l"] - chron[i-1]["pretax_eur_l"]
        db = chron[i - lag]["brent_eur_l"] - chron[i - lag - 1]["brent_eur_l"]
        if abs(db) >= threshold:
            d_retail.append(dr)
            d_brent.append(db)

    if not d_brent:
        return None

    up_pairs = [(r, b) for r, b in zip(d_retail, d_brent) if b > 0]
    dn_pairs = [(r, b) for r, b in zip(d_retail, d_brent) if b < 0]

    def mean_ratio(pairs):
        if len(pairs) < 3:
            return None
        return sum(r / b for r, b in pairs) / len(pairs)

    pt_up = mean_ratio(up_pairs)
    pt_dn = mean_ratio(dn_pairs)

    if pt_up is None or pt_dn is None:
        return None

    # Correlation (R²) between Δbrent and Δretail as fit quality indicator
    n = len(d_brent)
    mean_b = sum(d_brent) / n
    mean_r = sum(d_retail) / n
    ss_b = sum((b - mean_b) ** 2 for b in d_brent)
    ss_r = sum((r - mean_r) ** 2 for r in d_retail)
    sp   = sum((d_brent[i] - mean_b) * (d_retail[i] - mean_r) for i in range(n))
    r2   = (sp ** 2) / (ss_b * ss_r) if ss_b > 0 and ss_r > 0 else 0.0

    return {
        "lag":             lag,
        "pt_when_rising":  round(pt_up, 3),
        "pt_when_falling": round(pt_dn, 3),
        "asymmetry":       round(pt_up - pt_dn, 3),
        "r2":              round(r2, 3),
        "weeks_used":      n,
        "n_rising_weeks":  len(up_pairs),
        "n_falling_weeks": len(dn_pairs),
    }


def compute_pass_through(margin_history: list) -> dict:
    """
    Asymmetric pass-through analysis with lag detection.

    Tests lags 0–3 weeks and picks the lag with highest R² (best fit between
    Brent changes and retail responses). This detects whether retailers respond
    immediately or with a delay of 1–3 weeks.

    β_up:   mean ratio of Δ(pretax retail) / Δ(Brent EUR/L) for weeks when Brent rose
    β_down: mean ratio for weeks when Brent fell

    β_up > β_down → retailers pass on cost INCREASES faster than DECREASES
    (i.e. asymmetric — standard signal of market power / potential gouging)
    """
    chron = sorted(margin_history, key=lambda x: x["week"])
    if len(chron) < 8:
        return {"signal": "INSUFFICIENT_DATA", "weeks_used": len(chron)}

    # Test lags 0–3 weeks, pick the one with highest R²
    best = None
    all_lags = []
    for lag in range(4):
        result = _pass_through_core(chron, threshold=0.003, lag=lag)
        if result is not None:
            all_lags.append(result)
            if best is None or result["r2"] > best["r2"]:
                best = result

    if best is None:
        return {"signal": "INSUFFICIENT_DATA", "weeks_used": 0}

    asymmetry = best["asymmetry"]
    if asymmetry > 0.15:
        signal = "ASYMMETRIC"
    elif asymmetry < -0.15:
        signal = "FAVORABLE"
    else:
        signal = "SYMMETRIC"

    pt_up = best["pt_when_rising"]
    pt_dn = best["pt_when_falling"]

    return {
        "signal":          signal,
        "asymmetry":       asymmetry,
        "best_lag_weeks":  best["lag"],
        "lag_r2":          best["r2"],
        "pt_when_rising":  pt_up,
        "pt_when_falling": pt_dn,
        "weeks_used":      best["weeks_used"],
        "n_rising_weeks":  best["n_rising_weeks"],
        "n_falling_weeks": best["n_falling_weeks"],
        "all_lags":        all_lags,
        "interpretation": (
            f"Best fit at lag={best['lag']}w (R²={best['r2']:.3f}). "
            f"When Brent rises, retail moves {pt_up:.2f}× as much. "
            f"When Brent falls, retail moves {pt_dn:.2f}× as much."
        ),
    }


def compute_rolling_pass_through(margin_history: list, window: int = 26) -> list:
    """
    Compute pass-through stats over rolling windows to track how pricing
    behaviour evolves over time. Returns one entry per window-end week.
    """
    chron = sorted(margin_history, key=lambda x: x["week"])
    if len(chron) < window:
        return []

    results = []
    for end in range(window, len(chron) + 1):
        segment = chron[end - window:end]
        result = _pass_through_core(segment, threshold=0.003, lag=0)
        if result is None:
            continue
        results.append({
            "week_end":        segment[-1]["week"],
            "week_start":      segment[0]["week"],
            "pt_when_rising":  result["pt_when_rising"],
            "pt_when_falling": result["pt_when_falling"],
            "asymmetry":       result["asymmetry"],
            "r2":              result["r2"],
        })

    results.sort(key=lambda x: x["week_end"], reverse=True)
    return results


def compute_margin_stats(margin_history: list) -> dict:
    """
    Compute summary statistics and confidence bands for the gross margin series.
    """
    if not margin_history:
        return {}

    margins = [h["gross_margin"] for h in margin_history if h.get("gross_margin") is not None]
    if len(margins) < 4:
        return {}

    n    = len(margins)
    mean = sum(margins) / n
    var  = sum((m - mean) ** 2 for m in margins) / (n - 1)
    std  = var ** 0.5

    sorted_m = sorted(margins)
    q1 = sorted_m[n // 4]
    q3 = sorted_m[3 * n // 4]

    # Recent trend: last 12 weeks vs prior 12 weeks (filter out missing margins)
    recent = [h["gross_margin"] for h in margin_history[:12] if h.get("gross_margin") is not None]
    prior  = [h["gross_margin"] for h in margin_history[12:24] if h.get("gross_margin") is not None]
    recent_avg = sum(recent) / len(recent) if recent else None
    prior_avg  = sum(prior) / len(prior) if prior else None
    trend = None
    if recent_avg is not None and prior_avg is not None:
        diff = recent_avg - prior_avg
        if diff > 0.005:
            trend = "WIDENING"
        elif diff < -0.005:
            trend = "NARROWING"
        else:
            trend = "STABLE"

    return {
        "mean":       round(mean, 4),
        "std":        round(std, 4),
        "band_upper": round(mean + std, 4),
        "band_lower": round(mean - std, 4),
        "q1":         round(q1, 4),
        "q3":         round(q3, 4),
        "min":        round(sorted_m[0], 4),
        "max":        round(sorted_m[-1], 4),
        "n_weeks":    n,
        "trend":      trend,
        "recent_12w_avg": round(recent_avg, 4) if recent_avg is not None else None,
        "prior_12w_avg":  round(prior_avg, 4) if prior_avg is not None else None,
    }


def compute_seasonal_analysis(margin_history: list) -> dict:
    """
    Detect seasonal pricing patterns by grouping margins into heating season
    (Oct–Mar) vs off-season (Apr–Sep). Irish heating oil demand peaks in winter,
    so margins may widen seasonally.
    """
    if len(margin_history) < 26:
        return {}

    heating = []  # Oct–Mar
    off     = []  # Apr–Sep
    for h in margin_history:
        month = int(h["week"][5:7])
        gm = h.get("gross_margin")
        if gm is None:
            continue
        if month >= 10 or month <= 3:
            heating.append(gm)
        else:
            off.append(gm)

    if len(heating) < 4 or len(off) < 4:
        return {}

    heat_avg = sum(heating) / len(heating)
    off_avg  = sum(off) / len(off)
    premium  = heat_avg - off_avg

    return {
        "heating_season_avg_margin": round(heat_avg, 4),
        "off_season_avg_margin":     round(off_avg, 4),
        "seasonal_premium":          round(premium, 4),
        "heating_weeks":             len(heating),
        "off_season_weeks":          len(off),
        "signal": "WINTER_PREMIUM" if premium > 0.005 else
                  "SUMMER_PREMIUM" if premium < -0.005 else "NO_SEASONAL_EFFECT",
        "interpretation": (
            f"Heating season margin avg €{heat_avg:.4f}/L vs "
            f"off-season €{off_avg:.4f}/L "
            f"({'winter +' if premium > 0 else 'summer +'}{abs(premium)*100:.2f}c/L)."
        ),
    }


# ── Gouging detector ─────────────────────────────────────────────────────────

def compute_gouging_detector(margin_history: list) -> dict:
    """
    Compare actual retail price to a 'justified price' that assumes full
    pass-through of Brent cost changes. Weeks where retail exceeds the
    justified price significantly may indicate margin expansion / gouging.
    """
    if len(margin_history) < 12:
        return {}

    chron = sorted(margin_history, key=lambda x: x["week"])

    results = []
    for i in range(1, len(chron)):
        cur, prev = chron[i], chron[i - 1]
        d_brent = cur["brent_eur_l"] - prev["brent_eur_l"]
        justified = prev["pretax_eur_l"] + d_brent
        actual    = cur["pretax_eur_l"]
        gap       = actual - justified

        withtax = cur.get("withtax_eur_l")
        if withtax and cur["pretax_eur_l"] > 0:
            tax_mult = withtax / cur["pretax_eur_l"]
            justified_wt = round(justified * tax_mult, 4)
            gap_wt       = round(withtax - justified_wt, 4)
        else:
            justified_wt = gap_wt = None

        results.append({
            "week": cur["week"],
            "actual_pretax":    round(actual, 4),
            "justified_pretax": round(justified, 4),
            "gap_pretax":       round(gap, 4),
            "actual_withtax":   withtax,
            "justified_withtax": justified_wt,
            "gap_withtax":      gap_wt,
            "brent_change":     round(d_brent, 4),
        })

    results.sort(key=lambda x: x["week"], reverse=True)

    gaps = [r["gap_pretax"] for r in results]
    mean_gap = sum(gaps) / len(gaps)
    std_gap  = (sum((g - mean_gap) ** 2 for g in gaps) / (len(gaps) - 1)) ** 0.5 if len(gaps) > 1 else 0

    recent = results[:4] if len(results) >= 4 else results
    avg_recent = sum(r["gap_pretax"] for r in recent) / len(recent) if recent else 0

    if avg_recent > mean_gap + 2 * std_gap:
        signal = "ALERT"
    elif avg_recent > mean_gap + 1.5 * std_gap:
        signal = "ELEVATED"
    else:
        signal = "NORMAL"

    return {
        "signal":             signal,
        "recent_avg_gap":     round(avg_recent, 4),
        "historical_mean_gap": round(mean_gap, 4),
        "historical_std_gap": round(std_gap, 4),
        "threshold_elevated": round(mean_gap + 1.5 * std_gap, 4),
        "threshold_alert":    round(mean_gap + 2 * std_gap, 4),
        "history":            results[:52],
    }


# ── Carbon tax projection ────────────────────────────────────────────────────

def compute_carbon_projection() -> list:
    """
    Project carbon tax cost per litre and annual household cost to 2030.
    Ireland's carbon tax rises €7.50/tonne/year until €100/tonne in 2030.
    """
    schedule = [
        (2021, 33.50), (2022, 41.00), (2023, 48.50), (2024, 56.00),
        (2025, 63.50), (2026, 71.00), (2027, 78.50), (2028, 86.00),
        (2029, 93.50), (2030, 100.00),
    ]
    ANNUAL_L = 2000
    return [
        {
            "year":  y,
            "rate_per_tonne":           rate,
            "carbon_per_litre":         round(rate * CO2_FACTOR_KEROSENE, 4),
            "annual_carbon_cost_2000l": round(rate * CO2_FACTOR_KEROSENE * ANNUAL_L, 2),
        }
        for y, rate in schedule
    ]


# ── Main ──────────────────────────────────────────────────────────────────────

def main() -> None:
    now_utc = datetime.now(timezone.utc)
    print(f"[{now_utc.isoformat()}] Fetching EU Oil Bulletin...")

    existing: dict = {}
    if os.path.exists(OUTPUT_PATH):
        try:
            with open(OUTPUT_PATH) as f:
                existing = json.load(f)
        except Exception as e:
            print(f"  Warning: could not read existing eu_bulletin.json: {e}")

    # ── Load prices.json for Brent + EUR/USD ─────────────────────────────────
    brent_weekly: list = []
    eurusd_daily: list = []
    if os.path.exists(PRICES_PATH):
        try:
            with open(PRICES_PATH) as f:
                prices = json.load(f)
            brent_weekly = prices.get("brent_weekly", [])
            eurusd_daily = prices.get("eurusd", [])
            print(f"  Loaded prices.json: {len(brent_weekly)} weekly Brent pts, {len(eurusd_daily)} EUR/USD pts")
        except Exception as e:
            print(f"  Warning: could not read prices.json: {e}")
    else:
        print("  Warning: prices.json not found; margin analysis will be skipped")

    try:
        print("  Finding XLSX download link...")
        xlsx_url  = find_xlsx_url()
        print(f"  Downloading: {xlsx_url}")
        xlsx_data = download_xlsx(xlsx_url)
        print(f"  Downloaded {len(xlsx_data):,} bytes")

        print("  Parsing with-tax prices...")
        ireland_history, eu_comparison, as_of = parse_with_tax(xlsx_data)
        print(f"  With-tax: {len(ireland_history)} weeks  latest={as_of}  {ireland_history[0]['price_per_litre']:.4f} €/L")
        print(f"  EU comparison: {len(eu_comparison)} countries")

        print("  Parsing pre-tax (ex-tax) prices...")
        pretax_history = parse_pretax(xlsx_data)
        print(f"  Pre-tax: {len(pretax_history)} weeks  latest={pretax_history[0]['week']}  {pretax_history[0]['pretax_eur_l']:.4f} €/L")

        print("  Computing gross margin history...")
        margin_history = compute_margin_history(pretax_history, ireland_history, brent_weekly, eurusd_daily)
        print(f"  Margin history: {len(margin_history)} weeks")
        if margin_history:
            h0 = margin_history[0]
            print(f"  Latest ({h0['week']}): pretax={h0['pretax_eur_l']:.4f} brent={h0['brent_eur_l']:.4f} margin={h0['gross_margin']:.4f} €/L")

        print("  Computing pass-through stats (with lag detection)...")
        pt_stats = compute_pass_through(margin_history)
        print(f"  Pass-through signal: {pt_stats['signal']}  "
              f"β_up={pt_stats.get('pt_when_rising','?')}  β_down={pt_stats.get('pt_when_falling','?')}  "
              f"best_lag={pt_stats.get('best_lag_weeks','?')}w")

        print("  Computing rolling pass-through (26w windows)...")
        rolling_pt = compute_rolling_pass_through(margin_history, window=26)
        print(f"  Rolling pass-through: {len(rolling_pt)} data points")

        print("  Computing margin statistics & confidence bands...")
        margin_stats = compute_margin_stats(margin_history)
        if margin_stats:
            print(f"  Margin: mean={margin_stats['mean']:.4f} ±{margin_stats['std']:.4f} €/L  trend={margin_stats.get('trend','?')}")

        print("  Computing seasonal analysis...")
        seasonal = compute_seasonal_analysis(margin_history)
        if seasonal:
            print(f"  Seasonal: {seasonal['signal']}  premium={seasonal.get('seasonal_premium',0):.4f} €/L")

        print("  Computing gouging detector...")
        gouging = compute_gouging_detector(margin_history)
        if gouging:
            print(f"  Gouging signal: {gouging['signal']}  recent_gap={gouging.get('recent_avg_gap',0):.4f} €/L")

        print("  Computing carbon tax projection...")
        carbon_proj = compute_carbon_projection()
        print(f"  Carbon projection: {len(carbon_proj)} years through 2030")

        output = {
            "last_updated":    now_utc.isoformat(),
            "source":          "EU Weekly Oil Bulletin — European Commission",
            "as_of":           as_of,
            "ireland_history": ireland_history,
            "eu_comparison":   eu_comparison,
            "margin_history":  margin_history,
            "pass_through":    pt_stats,
            "rolling_pass_through": rolling_pt,
            "margin_stats":    margin_stats,
            "seasonal":        seasonal,
            "gouging":         gouging,
            "carbon_projection": carbon_proj,
        }

    except Exception as e:
        print(f"  ERROR: {e}", file=sys.stderr)
        import traceback; traceback.print_exc()
        if existing:
            print("  Falling back to existing data.")
            existing["last_updated"] = now_utc.isoformat()
            output = existing
        else:
            sys.exit(1)

    os.makedirs(os.path.dirname(OUTPUT_PATH), exist_ok=True)
    with open(OUTPUT_PATH, "w") as f:
        json.dump(output, f, indent=2)
    print(f"Saved → {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
